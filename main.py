'''通过package箱唛自动填写excel模版'''

from openpyxl import load_workbook
from pypdf import PdfReader


def extract_package_base_data():
    """
    read asin, address from package file.
    :return: a dictionary with data.
    """
    reader = PdfReader("package.pdf")
    lines=reader.pages[0].extract_text().splitlines()
    data_dict = dict()
    #parse asin
    #first 12 elements is FBA id
    data_dict["asin"] = lines[12][:12]
    #parse address
    #index 7 is warehouse id
    data_dict["warehouse id"] = lines[7]
    data_dict["address"] = (f'{lines[7]} - {lines[8]}, {lines[9]}, '
                            f'{lines[10]}')
    data_dict["city"] = lines[9].split(",")[0]
    data_dict["state"]=lines[9].split(",")[1].split()[0]
    data_dict["zip code"]=lines[9].split(",")[1].split()[1]

    #input to get reference id
    data_dict["reference id"] = input("reference id: ")

    return data_dict


def make_sku_dict():
    """
    make sku dict from package.pdf
    :return: a sku dict containing sku name and its quantity
    """
    reader = PdfReader("package.pdf")
    sku_dict = {}
    for num in range(len(reader.pages)):
        extract_list = reader.pages[num].extract_text().splitlines()
        target_sku = extract_list[-3]
        if target_sku not in sku_dict.keys():
            sku_dict[target_sku] = 1
        else:
            sku_dict[target_sku] += 1

    return sku_dict


def fill_cell(data_dict,sku_dict):
    """
    fill data cells.
    :param data_dict: a dict. a dict of asin,address,quantity.
    :param sku_dict: a dict. a dict of sku and its quantity
    :return: None
    """
    wb = load_workbook('130001.xlsx')
    ws = wb.active

    ws['B1'].value = data_dict["asin"]
    ws['D1'].value = data_dict["asin"]
    ws['B15'].value = data_dict["address"]
    ws['B8'].value=data_dict["warehouse id"]
    ws['B12'].value=data_dict["zip code"]
    ws['B13'].value=data_dict["state"]
    ws['B14'].value=data_dict["city"]

    #make a list of box id
    box_id_list=[]
    for i,v in enumerate(sku_dict.values()):
        print(v,'v')
        if i ==0:
            box_id_list.append(f"{data_dict['asin']}000001~{v}")
        else:
            last_num=int(box_id_list[i-1].split('~')[-1])+1
            next_num=last_num+v-1
            box_id_list.append(f"{data_dict['asin']}00000"
                               f"{str(last_num)}~{str(next_num)}")


    #fill box id and reference id
    for num in range(len(box_id_list)):
        ws[f'A{num+20}'].value=box_id_list[num]
        ws[f'B{num + 20}'].value = data_dict["reference id"]

    #fill box quantity
    for i,v in enumerate(sku_dict.values()):
        ws[f'J{i + 20}'].value=v


    #total quantity of boxes
    data_dict['quantity'] = sum(sku_dict.values())
    ws['B18'].value = data_dict["quantity"]

    file=(f'{data_dict["asin"]}-{data_dict["warehouse id"]}-新细亚箱单发票.xlsx')
    wb.save(file)

fill_cell(extract_package_base_data(), make_sku_dict())